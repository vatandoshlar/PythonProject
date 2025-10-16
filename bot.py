#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import json
import re
from datetime import datetime
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup, ReplyKeyboardMarkup, KeyboardButton, ReplyKeyboardRemove
from telegram.ext import Application, CommandHandler, MessageHandler, CallbackQueryHandler, ContextTypes, filters, \
    ConversationHandler
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Bot konfiguratsiyasi
BOT_TOKEN = os.getenv('BOT_TOKEN', '7729290828:AAFJl5pxtdnyvA6czTtcDQ3iexVq_Fd7_o0')
ADMIN_CHAT_ID = os.getenv('ADMIN_CHAT_ID', '7605860772')
GROUP_ID = os.getenv('GROUP_ID', '-1002930763309')

# Conversation states
(START_MENU, FULLNAME, COUNTRY, CITY, BIRTHDATE, PHONE, WORKPLACE,
 SPECIALTY, EDUCATION, NOMINATION, CREATIVE_WORK) = range(11)

# Ma'lumotlar fayli
DATA_FILE = 'registered_users.json'

# Global ma'lumotlar
registered_users = []

# Ma'lumot darajalari
EDUCATION_LEVELS = [
    "O'rta ta'lim",
    "O'rta maxsus ta'lim",
    "Oliy ta'lim (Bakalavr)",
    "Oliy ta'lim (Magistr)",
    "Oliy ta'lim (Doktorantura)",
    "Boshqa"
]

# Nominatsiyalar
NOMINATIONS = [
    "Eng tashabuskor o'zbek tili ustozi",
    "Eng yaxshi insho",
    "Eng yaxshi videorolik"
]


def load_data():
    global registered_users
    try:
        if os.path.exists(DATA_FILE):
            with open(DATA_FILE, 'r', encoding='utf-8') as f:
                registered_users = json.load(f)
        print("Ma'lumotlar yuklandi")
    except Exception as e:
        print(f"Xato: {e}")


def save_data():
    try:
        with open(DATA_FILE, 'w', encoding='utf-8') as f:
            json.dump(registered_users, f, ensure_ascii=False, indent=2)
        print("Ma'lumotlar saqlandi")
    except Exception as e:
        print(f"Xato: {e}")


def save_partial_user_data(user_id, context_data, step_name):
    """Save user data at each step, even if registration is incomplete"""
    try:
        # Check if user already exists in partial data
        existing_user = None
        for i, user in enumerate(registered_users):
            if user.get('user_id') == user_id:
                existing_user = i
                break
        
        # Create or update user data
        user_info = {
            'user_id': user_id,
            'username': context_data.get('username', "Username yo'q"),
            'first_name': context_data.get('first_name', ""),
            'last_name': context_data.get('last_name', ""),
            'language_code': context_data.get('language_code', ""),
            'last_updated': datetime.now().strftime('%d.%m.%Y %H:%M:%S'),
            'registration_status': 'incomplete',
            'current_step': step_name
        }
        
        # Add all available data from context
        for key, value in context_data.items():
            if key not in ['user_id', 'username', 'first_name', 'last_name', 'language_code']:
                user_info[key] = value
        
        if existing_user is not None:
            # Update existing user
            registered_users[existing_user] = user_info
        else:
            # Add new user
            registered_users.append(user_info)
        
        save_data()
        print(f"Partial data saved for user {user_id} at step: {step_name}")
    except Exception as e:
        print(f"Error saving partial data: {e}")


async def send_reminder_to_incomplete_users():
    """Send reminder messages to users with incomplete registrations"""
    try:
        incomplete_users = [u for u in registered_users if u.get('registration_status') == 'incomplete']
        
        if not incomplete_users:
            print("No incomplete registrations to remind")
            return
        
        print(f"Found {len(incomplete_users)} incomplete registrations")
        
        for user in incomplete_users:
            try:
                user_id = user.get('user_id')
                current_step = user.get('current_step', 'Noma\'lum')
                last_updated = user.get('last_updated', '')
                
                # Create reminder message based on current step
                step_messages = {
                    'fullname': "👤 Ismingizni kiriting",
                    'country': "🌍 Davlatni tanlang", 
                    'city': "🏙️ Shaharni kiriting",
                    'birthdate': "📅 Tug'ilgan sanangizni kiriting",
                    'phone': "📱 Telefon raqamingizni kiriting",
                    'workplace': "🏢 Ish joyingizni kiriting",
                    'specialty': "💼 Mutaxassisligingizni kiriting",
                    'education': "🎓 Ma'lumot darajangizni tanlang",
                    'nomination': "🏆 Nominatsiyani tanlang"
                }
                
                next_step = step_messages.get(current_step, "Ro'yxatdan o'tishni davom eting")
                
                reminder_text = (
                    f"⏰ <b>Eslatma!</b>\n\n"
                    f"📋 Sizning ro'yxatdan o'tish jarayoningiz to'liq yakunlanmagan.\n\n"
                    f"🔄 Keyingi qadam: {next_step}\n\n"
                    f"📅 Oxirgi yangilanish: {last_updated}\n\n"
                    f"✅ Ro'yxatdan o'tishni yakunlash uchun /start ni bosing va jarayonni davom eting.\n\n"
                    f"❓ Yordam kerak bo'lsa, admin bilan bog'laning: @L19671"
                )
                
                # Send reminder message
                await context.bot.send_message(
                    chat_id=user_id,
                    text=reminder_text,
                    parse_mode='HTML'
                )
                
                print(f"Reminder sent to user {user_id} (step: {current_step})")
                
            except Exception as e:
                print(f"Error sending reminder to user {user.get('user_id', 'unknown')}: {e}")
                
    except Exception as e:
        print(f"Error in send_reminder_to_incomplete_users: {e}")


async def handle_forwarded_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle forwarded messages from admin to save as new user"""
    user_id = str(update.effective_user.id)
    
    # Check if message is from admin
    if user_id != ADMIN_CHAT_ID:
        return
    
    # Check if message is forwarded
    if not update.message.forward_from or not update.message.forward_from_chat:
        return
    
    try:
        # Reload data to get latest users
        load_data()
        
        # Extract forwarded user info
        forwarded_user = update.message.forward_from
        forwarded_chat = update.message.forward_from_chat
        
        # Create new user entry from forwarded data
        new_user = {
            'user_id': forwarded_user.id,
            'username': forwarded_user.username or "Username yo'q",
            'first_name': forwarded_user.first_name or "",
            'last_name': forwarded_user.last_name or "",
            'language_code': forwarded_user.language_code or "",
            'registration_date': datetime.now().strftime('%d.%m.%Y %H:%M:%S'),
            'registration_status': 'complete',
            'completion_date': datetime.now().strftime('%d.%m.%Y %H:%M:%S'),
            'fullname': f"{forwarded_user.first_name or ''} {forwarded_user.last_name or ''}".strip(),
            'country': 'Noma\'lum',
            'city': 'Noma\'lum',
            'birthdate': 'Noma\'lum',
            'phone': 'Noma\'lum',
            'workplace': 'Noma\'lum',
            'specialty': 'Noma\'lum',
            'education': 'Noma\'lum',
            'nomination': 'Noma\'lum',
            'file': {
                'file_id': 'Noma\'lum',
                'file_type': 'Noma\'lum',
                'file_name': 'Noma\'lum'
            },
            'recovered_from_forward': True,
            'original_forward_date': datetime.now().strftime('%d.%m.%Y %H:%M:%S')
        }
        
        # Check if user already exists
        existing_user = None
        for i, user in enumerate(registered_users):
            if user.get('user_id') == forwarded_user.id:
                existing_user = i
                break
        
        if existing_user is not None:
            # Update existing user
            registered_users[existing_user] = new_user
            action = "yangilandi"
        else:
            # Add new user
            registered_users.append(new_user)
            action = "qo'shildi"
        
        save_data()
        
        # Send confirmation to admin
        await update.message.reply_text(
            f"✅ <b>Foydalanuvchi ma'lumotlari {action}!</b>\n\n"
            f"👤 Foydalanuvchi: {new_user['fullname']}\n"
            f"🆔 ID: {new_user['user_id']}\n"
            f"📅 Sana: {new_user['registration_date']}\n"
            f"🔄 Holat: Qayta tiklandi\n\n"
            f"📊 Jami foydalanuvchilar: {len(registered_users)}",
            parse_mode='HTML'
        )
        
        print(f"User {new_user['user_id']} {action} from forwarded message")
        
    except Exception as e:
        await update.message.reply_text(f"❌ Xato: {e}")
        print(f"Error handling forwarded message: {e}")


async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data.clear()

    text = (
        "👋 Assalomu alaykum!\n\n"
        "🎊 Fondning tanlov botiga xush kelibsiz!\n\n"
        "📋 Iltimos, tanlovda qatnashishdan oldin shartlarni Nizom orqali tanishib chiqishingizni so'raymiz. \n\n"
        "Tanlovda omad tilaymiz! 🍀"
    )

    keyboard = [
        [
            InlineKeyboardButton("📋 Tanlov shartlari bilan tanishish", url="https://api.vatandoshlarfondi.uz/storage/maple-users/October2025/CD44Dg7oUTAwAhRH32uN.pdf")
        ],
        [
            InlineKeyboardButton("✅ Ro'yxatdan o'tishni boshlash", callback_data="begin_reg")
        ]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)

    if update.message:
        await update.message.reply_text(text, reply_markup=reply_markup)
    else:
        # In case /start is triggered from a callback or other update types
        await context.bot.send_message(chat_id=update.effective_chat.id, text=text, reply_markup=reply_markup)

    return START_MENU


async def begin_registration_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    await context.bot.send_chat_action(chat_id=update.effective_chat.id, action="typing")
    await context.bot.send_message(
        chat_id=update.effective_chat.id,
        text="✅",
        reply_markup=ReplyKeyboardRemove()
    )
    await context.bot.send_message(
        chat_id=update.effective_chat.id,
        text=(
            "Ro'yxatdan o'tish uchun quyidagi ma'lumotlarni to'ldiring.\n\n"
            "📝 Iltimos, ismingiz, familiyangiz va sharifingizni kiriting:\n"
            "(Masalan: Ibragimov Samandar Iskandar o'g'li)"
        )
    )

    return FULLNAME


async def handle_broadcast_qatnashish(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Before START_MENU: any message (including button) shows the Start menu. Keyboard remains."""
    return await start(update, context)


async def startmenu_catch_all(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """At Start menu: any message immediately starts registration and removes keyboard."""
    if update.message:
        await update.message.chat.send_action(action="typing")
        await update.message.reply_text("✅", reply_markup=ReplyKeyboardRemove())
        await update.message.reply_text(
            "Ro'yxatdan o'tish uchun quyidagi ma'lumotlarni to'ldiring.\n\n"
            "📝 Iltimos, ismingiz, familiyangiz va sharifingizni kiriting:\n"
            "(Masalan: Ibragimov Samandar Iskandar o'g'li)"
        )
    
    return FULLNAME


async def fullname(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.message.document or update.message.photo or update.message.video or update.message.audio or update.message.voice or update.message.sticker:
        await update.message.reply_text(
            "❌ Iltimos, faqat <b>matn</b> kiriting!\n\n"
            "📝 Ismingiz, familiyangiz va sharifingizni matn ko'rinishida yuboring.\n"
            "(Masalan: Ibragimov Samandar Iskandar o'g'li)",
            parse_mode='HTML'
        )
        return FULLNAME

    if not update.message.text:
        await update.message.reply_text(
            "❌ Iltimos, <b>matn</b> yuboring!\n\n"
            "📝 Ismingiz, familiyangiz va sharifingizni kiriting.",
            parse_mode='HTML'
        )
        return FULLNAME

    context.user_data['fullname'] = update.message.text
    
    # Save partial data
    user = update.effective_user
    context.user_data.update({
        'user_id': user.id,
        'username': user.username or "Username yo'q",
        'first_name': user.first_name or "",
        'last_name': user.last_name or "",
        'language_code': user.language_code or ""
    })
    save_partial_user_data(user.id, context.user_data, 'fullname')
    
    await update.message.reply_text("🌍 Hozirda qaysi davlatda istiqomat qilasiz?")
    return COUNTRY


async def country(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.message.document or update.message.photo or update.message.video or update.message.audio or update.message.voice or update.message.sticker:
        await update.message.reply_text(
            "❌ Iltimos, faqat <b>matn</b> kiriting!\n\n"
            "🌍 Davlat nomini matn ko'rinishida yuboring.",
            parse_mode='HTML'
        )
        return COUNTRY

    if not update.message.text:
        await update.message.reply_text(
            "❌ Iltimos, <b>matn</b> yuboring!\n\n"
            "🌍 Davlat nomini kiriting.",
            parse_mode='HTML'
        )
        return COUNTRY

    context.user_data['country'] = update.message.text
    
    # Save partial data
    user = update.effective_user
    save_partial_user_data(user.id, context.user_data, 'country')
    
    await update.message.reply_text("🏙️ Istiqomat qilayotgan shahar yoki tumaningizni kiriting:")
    return CITY


async def city(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.message.document or update.message.photo or update.message.video or update.message.audio or update.message.voice or update.message.sticker:
        await update.message.reply_text(
            "❌ Iltimos, faqat <b>matn</b> kiriting!\n\n"
            "🏙️ Shahar yoki tuman nomini matn ko'rinishida yuboring.",
            parse_mode='HTML'
        )
        return CITY

    if not update.message.text:
        await update.message.reply_text(
            "❌ Iltimos, <b>matn</b> yuboring!\n\n"
            "🏙️ Shahar yoki tuman nomini kiriting.",
            parse_mode='HTML'
        )
        return CITY

    context.user_data['city'] = update.message.text
    
    # Save partial data
    user = update.effective_user
    save_partial_user_data(user.id, context.user_data, 'city')
    
    await update.message.reply_text(
        "📅 Tug'ilgan sanangizni kiriting (dd.mm.yyyy formatida):\n"
        "(Masalan: 04.06.1994)"
    )
    return BIRTHDATE


async def birthdate(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.message.document or update.message.photo or update.message.video or update.message.audio or update.message.voice or update.message.sticker:
        await update.message.reply_text(
            "❌ Iltimos, faqat <b>matn</b> kiriting!\n\n"
            "📅 Tug'ilgan sanangizni matn ko'rinishida yuboring (dd.mm.yyyy formatida).\n"
            "(Masalan: 04.06.1994)",
            parse_mode='HTML'
        )
        return BIRTHDATE

    if not update.message.text:
        await update.message.reply_text(
            "❌ Iltimos, <b>matn</b> yuboring!\n\n"
            "📅 Tug'ilgan sanangizni kiriting (dd.mm.yyyy formatida).",
            parse_mode='HTML'
        )
        return BIRTHDATE

    text = update.message.text

    try:
        datetime.strptime(text, '%d.%m.%Y')
        context.user_data['birthdate'] = text
        
        # Save partial data
        user = update.effective_user
        save_partial_user_data(user.id, context.user_data, 'birthdate')
        
        await update.message.reply_text(
            "📱 Telefon raqamingizni kiriting."
        )
        return PHONE
    except ValueError:
        await update.message.reply_text(
            "❌ Noto'g'ri format! Iltimos, dd.mm.yyyy formatida kiriting.\n"
            "(Masalan: 04.06.1994)"
        )
        return BIRTHDATE


async def phone(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.message.document or update.message.photo or update.message.video or update.message.audio or update.message.voice or update.message.sticker:
        await update.message.reply_text(
            "❌ Iltimos, faqat <b>matn</b> kiriting!\n\n"
            "📱 Telefon raqamingizni matn ko'rinishida yuboring.",
            parse_mode='HTML'
        )
        return PHONE

    if not update.message.text:
        await update.message.reply_text(
            "❌ Iltimos, <b>matn</b> yuboring!\n\n"
            "📱 Telefon raqamingizni kiriting.",
            parse_mode='HTML'
        )
        return PHONE

    phone_text = update.message.text.strip()

    context.user_data['phone'] = phone_text
    
    # Save partial data
    user = update.effective_user
    save_partial_user_data(user.id, context.user_data, 'phone')
    
    await update.message.reply_text("🏢 Ish yoki o'qish joyingizni kiriting:")
    return WORKPLACE


async def workplace(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.message.document or update.message.photo or update.message.video or update.message.audio or update.message.voice or update.message.sticker:
        await update.message.reply_text(
            "❌ Iltimos, faqat <b>matn</b> kiriting!\n\n"
            "🏢 Ish yoki o'qish joyingizni matn ko'rinishida yuboring.",
            parse_mode='HTML'
        )
        return WORKPLACE

    if not update.message.text:
        await update.message.reply_text(
            "❌ Iltimos, <b>matn</b> yuboring!\n\n"
            "🏢 Ish yoki o'qish joyingizni kiriting.",
            parse_mode='HTML'
        )
        return WORKPLACE

    context.user_data['workplace'] = update.message.text
    
    # Save partial data
    user = update.effective_user
    save_partial_user_data(user.id, context.user_data, 'workplace')
    
    await update.message.reply_text("💼 Mutaxassisligingizni kiriting:")
    return SPECIALTY


async def specialty(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.message.document or update.message.photo or update.message.video or update.message.audio or update.message.voice or update.message.sticker:
        await update.message.reply_text(
            "❌ Iltimos, faqat <b>matn</b> kiriting!\n\n"
            "💼 Mutaxassisligingizni matn ko'rinishida yuboring.",
            parse_mode='HTML'
        )
        return SPECIALTY

    if not update.message.text:
        await update.message.reply_text(
            "❌ Iltimos, <b>matn</b> yuboring!\n\n"
            "💼 Mutaxassisligingizni kiriting.",
            parse_mode='HTML'
        )
        return SPECIALTY

    context.user_data['specialty'] = update.message.text
    
    # Save partial data
    user = update.effective_user
    save_partial_user_data(user.id, context.user_data, 'specialty')

    keyboard = [[InlineKeyboardButton(level, callback_data=f"edu_{i}")]
                for i, level in enumerate(EDUCATION_LEVELS)]
    reply_markup = InlineKeyboardMarkup(keyboard)

    await update.message.reply_text(
        "🎓 Ma'lumot darajangizni tanlang:",
        reply_markup=reply_markup
    )
    return EDUCATION


async def education_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    edu_index = int(query.data.split('_')[1])
    context.user_data['education'] = EDUCATION_LEVELS[edu_index]
    
    # Save partial data
    user = update.effective_user
    save_partial_user_data(user.id, context.user_data, 'education')

    await query.edit_message_text(f"✅ Tanlandi: {EDUCATION_LEVELS[edu_index]}")

    keyboard = [[InlineKeyboardButton(nom, callback_data=f"nom_{i}")]
                for i, nom in enumerate(NOMINATIONS)]
    reply_markup = InlineKeyboardMarkup(keyboard)

    await context.bot.send_message(
        chat_id=update.effective_chat.id,
        text="🏆 Nominatsiyani tanlang:",
        reply_markup=reply_markup
    )
    return NOMINATION


async def education_invalid_input(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "❌ Iltimos, yuqoridagi tugmalardan birini bosing!\n\n"
        "🎓 Ma'lumot darajangizni tugmalar orqali tanlang."
    )
    return EDUCATION


async def nomination_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    nom_index = int(query.data.split('_')[1])
    context.user_data['nomination'] = NOMINATIONS[nom_index]
    
    # Save partial data
    user = update.effective_user
    save_partial_user_data(user.id, context.user_data, 'nomination')

    await query.edit_message_text(f"✅ Tanlandi: {NOMINATIONS[nom_index]}")

    await context.bot.send_message(
        chat_id=update.effective_chat.id,
        text="📎 Ijodiy ishingizni yuboring:\n(Video, audio, PDF, PowerPoint yoki boshqa hujjat)"
    )
    return CREATIVE_WORK


async def nomination_invalid_input(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "❌ Iltimos, yuqoridagi tugmalardan birini bosing!\n\n"
        "🏆 Nominatsiyani tugmalar orqali tanlang."
    )
    return NOMINATION


async def creative_work(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user

    if update.message.text:
        await update.message.reply_text(
            "❌ Iltimos, matn emas, <b>fayl</b> yuboring!\n\n"
            "📎 Quyidagi fayl turlarini yuborishingiz mumkin:\n"
            "• 🎥 Video fayl\n"
            "• 🎵 Audio fayl\n"
            "• 📄 PDF hujjat\n"
            "• 📊 PowerPoint (pptx)\n"
            "• 📝 Word hujjat\n"
            "• 📁 Boshqa hujjatlar\n\n"
            "🔄 Iltimos, fayl yuboring.",
            parse_mode='HTML'
        )
        return CREATIVE_WORK

    if update.message.photo:
        await update.message.reply_text(
            "❌ Rasm emas, iltimos to'g'ri <b>fayl</b> yuboring!\n\n"
            "📎 Quyidagi fayl turlarini yuborishingiz mumkin:\n"
            "• 🎥 Video fayl\n"
            "• 🎵 Audio fayl\n"
            "• 📄 PDF hujjat\n"
            "• 📊 PowerPoint (pptx)\n"
            "• 📝 Word hujjat\n"
            "• 📁 Boshqa hujjatlar\n\n"
            "🔄 Iltimos, qaytadan urinib ko'ring.",
            parse_mode='HTML'
        )
        return CREATIVE_WORK

    if not (update.message.document or update.message.video or update.message.audio):
        await update.message.reply_text(
            "❌ Iltimos, to'g'ri fayl turini yuboring!\n\n"
            "📎 Quyidagi fayl turlarini yuborishingiz mumkin:\n"
            "• 🎥 Video fayl\n"
            "• 🎵 Audio fayl\n"
            "• 📄 PDF hujjat\n"
            "• 📊 PowerPoint (pptx)\n"
            "• 📝 Word hujjat\n"
            "• 📁 Boshqa hujjatlar\n\n"
            "🔄 Iltimos, qaytadan urinib ko'ring.",
            parse_mode='HTML'
        )
        return CREATIVE_WORK

    file_info = {}

    if update.message.document:
        file_info = {
            'file_id': update.message.document.file_id,
            'file_type': 'document',
            'file_name': update.message.document.file_name
        }
    elif update.message.video:
        file_info = {
            'file_id': update.message.video.file_id,
            'file_type': 'video',
            'file_name': 'video'
        }
    elif update.message.audio:
        file_info = {
            'file_id': update.message.audio.file_id,
            'file_type': 'audio',
            'file_name': 'audio'
        }

    context.user_data['file'] = file_info

    # Remove any existing partial data for this user
    registered_users[:] = [u for u in registered_users if u.get('user_id') != user.id]

    user_info = {
        'user_id': user.id,
        'username': user.username or "Username yo'q",
        'first_name': user.first_name or "",
        'last_name': user.last_name or "",
        'language_code': user.language_code or "",
        'registration_date': datetime.now().strftime('%d.%m.%Y %H:%M:%S'),
        'registration_status': 'complete',
        'completion_date': datetime.now().strftime('%d.%m.%Y %H:%M:%S')
    }
    user_info.update(context.user_data)

    registered_users.append(user_info)
    save_data()

    await update.message.reply_text(
        "✅ Ro'yxatdan muvaffaqiyatli o'tdingiz!\n\n"
        "📋 Sizning ma'lumotlaringiz:\n"
        f"👤 F.I.Sh: {user_info['fullname']}\n"
        f"🌍 Davlat: {user_info['country']}\n"
        f"🏙️ Shahar/Tuman: {user_info['city']}\n"
        f"📅 Tug'ilgan sana: {user_info['birthdate']}\n"
        f"📱 Telefon: {user_info['phone']}\n"
        f"🏢 Ish joyi: {user_info['workplace']}\n"
        f"💼 Mutaxassislik: {user_info['specialty']}\n"
        f"🎓 Ma'lumot: {user_info['education']}\n"
        f"🏆 Nominatsiya: {user_info['nomination']}"
    )

    await update.message.reply_text("📎 Sizning ijodiy ishingiz:")

    if file_info['file_type'] == 'video':
        await context.bot.send_video(
            chat_id=update.effective_chat.id,
            video=file_info['file_id'],
            caption=f"🎥 Video fayl: {file_info['file_name']}"
        )
    elif file_info['file_type'] == 'audio':
        await context.bot.send_audio(
            chat_id=update.effective_chat.id,
            audio=file_info['file_id'],
            caption=f"🎵 Audio fayl: {file_info['file_name']}"
        )
    else:
        await context.bot.send_document(
            chat_id=update.effective_chat.id,
            document=file_info['file_id'],
            caption=f"📄 Hujjat: {file_info['file_name']}"
        )

    # Show typing while sending to admin
    await update.message.chat.send_action(action="typing")
    await send_to_admin(context, user_info)
    
    await update.message.reply_text(
        "✅ Ma'lumotlaringiz administratorga yuborildi.\n"
        "Rahmat! 🙏"
    )
    return ConversationHandler.END


async def send_to_admin(context: ContextTypes.DEFAULT_TYPE, user_info: dict):
    print(f"GROUP_ID ishlatilmoqda: {GROUP_ID}")
    print(f"ADMIN_CHAT_ID ishlatilmoqda: {ADMIN_CHAT_ID}")

    try:
        extra_info = ""
        if user_info.get('first_name'):
            extra_info += f"📝 <b>Ism (Telegram):</b> {user_info.get('first_name', '')}\n"
        if user_info.get('last_name'):
            extra_info += f"📝 <b>Familiya (Telegram):</b> {user_info.get('last_name', '')}\n"
        if user_info.get('language_code'):
            extra_info += f"🌐 <b>Til:</b> {user_info.get('language_code', '')}\n"

        message = (
            "📋 <b>Yangi ro'yxatdan o'tish</b>\n\n"
            f"👤 <b>F.I.Sh:</b> {user_info['fullname']}\n"
            f"🌍 <b>Davlat:</b> {user_info['country']}\n"
            f"🏙️ <b>Shahar/Tuman:</b> {user_info['city']}\n"
            f"🎂 <b>Tug'ilgan sana:</b> {user_info['birthdate']}\n"
            f"📱 <b>Telefon:</b> {user_info['phone']}\n"
            f"📞 <b>Telegram:</b> @{user_info['username']}\n"
            f"{extra_info}"
            f"🏢 <b>Ish joyi:</b> {user_info['workplace']}\n"
            f"💼 <b>Mutaxassislik:</b> {user_info['specialty']}\n"
            f"🎓 <b>Ma'lumot:</b> {user_info['education']}\n"
            f"🏆 <b>Nominatsiya:</b> {user_info['nomination']}\n\n"
            f"⏰ <b>Ro'yxatdan o'tgan vaqt:</b> {user_info['registration_date']}"
        )

        file_info = user_info['file']

        keyboard = [
            [
                InlineKeyboardButton("✅ Tasdiqlash", callback_data=f"approve_{user_info['user_id']}"),
                InlineKeyboardButton("❌ Rad etish", callback_data=f"reject_{user_info['user_id']}")
            ]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)

        sent_message = None

        if file_info['file_type'] == 'video':
            sent_message = await context.bot.send_video(
                chat_id=GROUP_ID,
                video=file_info['file_id'],
                caption=message,
                parse_mode='HTML',
                reply_markup=reply_markup
            )
        elif file_info['file_type'] == 'audio':
            sent_message = await context.bot.send_audio(
                chat_id=GROUP_ID,
                audio=file_info['file_id'],
                caption=message,
                parse_mode='HTML',
                reply_markup=reply_markup
            )
        else:
            sent_message = await context.bot.send_document(
                chat_id=GROUP_ID,
                document=file_info['file_id'],
                caption=message,
                parse_mode='HTML',
                reply_markup=reply_markup
            )

        if sent_message:
            group_id_str = str(GROUP_ID).replace('-100', '')
            message_link = f"https://t.me/c/{group_id_str}/{sent_message.message_id}"

            for user in registered_users:
                if user['user_id'] == user_info['user_id']:
                    user['message_link'] = message_link
                    user['message_id'] = sent_message.message_id
                    break

            save_data()

        print("Adminga yuborildi")
    except Exception as e:
        print(f"Admin xato: {e}")
        import traceback
        traceback.print_exc()


async def broadcast_previous_users(application: Application):
    """Send a broadcast message to all previously registered users (by user_id)."""
    try:
        # Reload data to get latest users
        load_data()
        if not registered_users:
            return

        seen = set()
        text = (
            "<b>👋 Assalomu alaykum!</b>\n\n"
            "<b>⚡️ Vatandoshlar jamoat fondi ”Sen turfa jilosan aziz ona til” nomli yangi tanlov e'lon qildi.</b>\n\n"
            "<b>✨ Qatnashish va batafsil ma'lumot olish uchun /start yoki qatnashish tugmasini bosing</b>"
            "<em>\n\n⚠ Ushbu xabar hammaga avtomatik tarzda yuborildi. Agar siz ro'yxatdan o'tib bo'lgan bo'lsangiz, ushbu xabarni e'tiborsiz qoldiring.</em>"
        )
        reply_kb = ReplyKeyboardMarkup([[KeyboardButton("✅ Qatnashish")]], resize_keyboard=True)

        success = 0
        for user in registered_users:
            user_id = user.get('user_id')
            if not user_id or user_id in seen:
                continue
            seen.add(user_id)
            try:
                await application.bot.send_message(chat_id=user_id, text=text, reply_markup=reply_kb, parse_mode='HTML')
                success += 1
            except Exception as send_err:
                print(f"Broadcast failed for {user_id}: {send_err}")
        return success
    except Exception as e:
        print(f"Broadcast error: {e}")


async def broadcast_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Admin-only manual trigger to broadcast message to previous users."""
    user_id = str(update.effective_user.id)
    if user_id != ADMIN_CHAT_ID:
        await update.message.reply_text("❌ Sizda ruxsat yo'q!")
        return
    
    total_users = len(set(u.get('user_id') for u in registered_users if u.get('user_id')))
    
    # Animated progress bar
    def get_progress_bar(current, total, length=10):
        filled = int((current / total) * length) if total > 0 else 0
        bar = "█" * filled + "░" * (length - filled)
        percent = int((current / total) * 100) if total > 0 else 0
        return f"{bar} {percent}%"
    
    progress_msg = await update.message.reply_text(
        f"🚀 <b>Broadcast boshlanmoqda...</b>\n\n"
        f"{get_progress_bar(0, total_users)}\n"
        f"Yuborildi: 0 / {total_users}",
        parse_mode='HTML'
    )
    
    # Track progress
    sent = 0
    seen = set()
    text = (
        "<b>👋 Assalomu alaykum!</b>\n\n"
        "<b>⚡️ Vatandoshlar jamoat fondi ”Sen turfa jilosan aziz ona til” nomli yangi tanlov e'lon qildi.</b>\n\n"
        "<b>✨ Qatnashish va batafsil ma'lumot olish uchun /start yoki qatnashish tugmasini bosing</b>"
        "<em>\n\n⚠ Ushbu xabar hammaga avtomatik tarzda yuborildi. Agar siz ro'yxatdan o'tib bo'lgan bo'lsangiz, ushbu xabarni e'tiborsiz qoldiring.</em>"
    )
    reply_kb = ReplyKeyboardMarkup([[KeyboardButton("✅ Qatnashish")]], resize_keyboard=True)
    
    for user in registered_users:
        user_id_val = user.get('user_id')
        if not user_id_val or user_id_val in seen:
            continue
        seen.add(user_id_val)
        try:
            await context.application.bot.send_message(chat_id=user_id_val, text=text, reply_markup=reply_kb, parse_mode='HTML')
            sent += 1
            # Update progress every 3 users with animated bar
            if sent % 3 == 0 or sent == total_users:
                await progress_msg.edit_text(
                    f"🚀 <b>Broadcast boshlanmoqda...</b>\n\n"
                    f"{get_progress_bar(sent, total_users)}\n"
                    f"Yuborildi: {sent} / {total_users}",
                    parse_mode='HTML'
                )
        except Exception as send_err:
            print(f"Broadcast failed for {user_id_val}: {send_err}")
    
    await progress_msg.edit_text(
        f"✅ <b>Broadcast tugadi!</b>\n\n"
        f"{get_progress_bar(sent, total_users)}\n"
        f"Jami yuborildi: <b>{sent}</b> ta foydalanuvchi\n"
        f"Xatoliklar: <b>{total_users - sent}</b> ta",
        parse_mode='HTML'
    )


async def approve_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer("✅ Tasdiqlandi!")
    await query.edit_message_reply_markup(reply_markup=None)


async def reject_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer("❌ Rad etildi!")
    await query.edit_message_reply_markup(reply_markup=None)


async def export_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = str(update.effective_user.id)
    print(f"Export so'raldi. User ID: {user_id}")
    print(f"Admin ID: {ADMIN_CHAT_ID}")

    if user_id != ADMIN_CHAT_ID:
        await update.message.reply_text(f"❌ Sizda ruxsat yo'q!")
        return

    # Reload data from file to get latest users
    load_data()
    print(f"Ro'yxatdan o'tganlar soni: {len(registered_users)}")

    if not registered_users:
        await update.message.reply_text("📊 Hozircha ro'yxatdan o'tgan foydalanuvchilar yo'q.")
        return

    try:
        await update.message.chat.send_action(action="upload_document")
        progress_msg = await update.message.reply_text(
            "📊 <b>Excel tayyorlanmoqda...</b>\n\n"
            "▱▱▱▱▱▱▱▱▱▱ 0%",
            parse_mode='HTML'
        )
        print("Excel yaratish boshlandi...")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Ro'yxatdan o'tganlar"

        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        headers = [
            "№", "Telegram ID", "F.I.Sh", "Davlat", "Shahar/Tuman", "Tug'ilgan sana", "Telefon",
            "Ish joyi", "Mutaxassislik", "Ma'lumot", "Nominatsiya",
            "Telegram", "Holat", "Oxirgi yangilanish", "Ijodiy ish"
        ]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border

        print("Sarlavhalar yozildi...")
        
        await progress_msg.edit_text(
            "📊 <b>Excel tayyorlanmoqda...</b>\n\n"
            "███▱▱▱▱▱▱▱ 30%\nMa'lumotlar yozilmoqda...",
            parse_mode='HTML'
        )

        for row_num, user in enumerate(registered_users, 2):
            work_link = user.get('message_link', 'Link yo\'q')
            
            # Determine status and date
            status = user.get('registration_status', 'complete')
            unknown_text = 'Noma\'lum'
            status_text = "✅ To'liq" if status == 'complete' else f"⏳ {user.get('current_step', unknown_text)}"
            date_field = user.get('completion_date', user.get('last_updated', user.get('registration_date', '')))

            data = [
                row_num - 1,
                user.get('user_id', ''),
                user.get('fullname', ''),
                user.get('country', ''),
                user.get('city', ''),
                user.get('birthdate', ''),
                user.get('phone', ''),
                user.get('workplace', ''),
                user.get('specialty', ''),
                user.get('education', ''),
                user.get('nomination', ''),
                f"@{user.get('username', '')}",
                status_text,
                date_field,
                work_link
            ]

            for col_num, value in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                cell.border = thin_border

                if col_num == 13 and work_link != 'Link yo\'q':
                    cell.hyperlink = work_link
                    cell.style = "Hyperlink"

        print("Ma'lumotlar yozildi...")
        
        await progress_msg.edit_text(
            "📊 <b>Excel tayyorlanmoqda...</b>\n\n"
            "███████▱▱▱ 70%\nFormatlash...",
            parse_mode='HTML'
        )

        column_widths = [5, 12, 25, 15, 15, 15, 15, 25, 20, 25, 30, 15, 20, 50]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = width

        ws.row_dimensions[1].height = 30
        for row_num in range(2, len(registered_users) + 2):
            ws.row_dimensions[row_num].height = 25

        excel_filename = f"royxat_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb.save(excel_filename)
        print(f"Excel fayl saqlandi: {excel_filename}")
        
        await progress_msg.edit_text(
            "📊 <b>Excel tayyorlanmoqda...</b>\n\n"
            "██████████ 100%\nYuborilmoqda...",
            parse_mode='HTML'
        )

        # Count complete and incomplete registrations
        complete_count = len([u for u in registered_users if u.get('registration_status') == 'complete'])
        incomplete_count = len([u for u in registered_users if u.get('registration_status') == 'incomplete'])
        
        with open(excel_filename, 'rb') as excel_file:
            await update.message.reply_document(
                document=excel_file,
                caption=f"📊 Jami {len(registered_users)} ta foydalanuvchi\n"
                       f"✅ To'liq: {complete_count} ta\n"
                       f"⏳ To'liq emas: {incomplete_count} ta\n\n"
                       f"📋 Excel formatda",
                filename=excel_filename
            )

        await progress_msg.delete()
        print("Excel yuborildi!")
        os.remove(excel_filename)
        print("Excel fayl o'chirildi")

        export_text = "📊 RO'YXATDAN O'TGAN FOYDALANUVCHILAR\n"
        export_text += "=" * 40 + "\n\n"

        for i, user in enumerate(registered_users, 1):
            work_link = user.get('message_link', 'Link yo\'q')

            export_text += f"{i}. 👤 {user.get('fullname', '')}\n"
            export_text += f"   🌍 Davlat: {user.get('country', '')}\n"
            export_text += f"   🏙️ Shahar/Tuman: {user.get('city', '')}\n"
            export_text += f"   📅 Tug'ilgan sana: {user.get('birthdate', '')}\n"
            export_text += f"   📱 Telefon: {user.get('phone', '')}\n"
            export_text += f"   🏢 Ish joyi: {user.get('workplace', '')}\n"
            export_text += f"   💼 Mutaxassislik: {user.get('specialty', '')}\n"
            export_text += f"   🎓 Ma'lumot: {user.get('education', '')}\n"
            export_text += f"   🏆 Nominatsiya: {user.get('nomination', '')}\n"
            export_text += f"   📞 Telegram: @{user.get('username', '')}\n"
            export_text += f"   ⏰ Vaqt: {user.get('registration_date', user.get('last_updated', ''))}\n"
            export_text += f"   📎 Ijodiy ish: {work_link}\n"
            export_text += "   " + "-" * 35 + "\n\n"

        filename = f"royxat_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
        with open(filename, 'w', encoding='utf-8') as f:
            f.write(export_text)

        with open(filename, 'rb') as txt_file:
            await update.message.reply_document(
                document=txt_file,
                caption="📄 Text formatda",
                filename=filename
            )

        print("Text yuborildi!")
        os.remove(filename)

        print("✅ Export muvaffaqiyatli tugadi!")

    except Exception as e:
        print(f"❌ Export xatosi: {e}")
        import traceback
        traceback.print_exc()
        await update.message.reply_text(f"❌ Xatolik yuz berdi:\n{str(e)}")


async def get_chat_id(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    chat_type = update.effective_chat.type
    chat_title = update.effective_chat.title or "Private chat"

    await update.message.reply_text(
        f"📊 Chat ma'lumotlari:\n\n"
        f"🆔 Chat ID: <code>{chat_id}</code>\n"
        f"📝 Chat type: {chat_type}\n"
        f"📌 Chat nomi: {chat_title}\n\n"
        f"ID ni nusxalash uchun bosing ☝️",
        parse_mode='HTML'
    )


async def reminder_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Admin command to send reminders to users with incomplete registrations"""
    user_id = str(update.effective_user.id)
    print(f"Reminder so'raldi. User ID: {user_id}")
    print(f"Admin ID: {ADMIN_CHAT_ID}")

    if user_id != ADMIN_CHAT_ID:
        await update.message.reply_text(f"❌ Sizda ruxsat yo'q!")
        return

    try:
        # Reload data from file to get latest users
        load_data()
        incomplete_users = [u for u in registered_users if u.get('registration_status') == 'incomplete']
        
        if not incomplete_users:
            await update.message.reply_text("📊 To'liq bo'lmagan ro'yxatdan o'tishlar yo'q.")
            return

        await update.message.reply_text(
            f"⏰ <b>Eslatma yuborilmoqda...</b>\n\n"
            f"📊 Jami {len(incomplete_users)} ta to'liq bo'lmagan ro'yxatdan o'tish\n\n"
            f"▱▱▱▱▱▱▱▱▱▱ 0%",
            parse_mode='HTML'
        )

        sent_count = 0
        failed_count = 0

        for i, user in enumerate(incomplete_users):
            try:
                user_id = user.get('user_id')
                current_step = user.get('current_step', 'Noma\'lum')
                last_updated = user.get('last_updated', '')
                
                # Create reminder message based on current step
                step_messages = {
                    'fullname': "👤 Ismingizni kiriting",
                    'country': "🌍 Davlatni tanlang", 
                    'city': "🏙️ Shaharni kiriting",
                    'birthdate': "📅 Tug'ilgan sanangizni kiriting",
                    'phone': "📱 Telefon raqamingizni kiriting",
                    'workplace': "🏢 Ish joyingizni kiriting",
                    'specialty': "💼 Mutaxassisligingizni kiriting",
                    'education': "🎓 Ma'lumot darajangizni tanlang",
                    'nomination': "🏆 Nominatsiyani tanlang"
                }
                
                next_step = step_messages.get(current_step, "Ro'yxatdan o'tishni davom eting")
                
                reminder_text = (
                    f"⏰ <b>Eslatma!</b>\n\n"
                    f"📋 Sizning ro'yxatdan o'tish jarayoningiz to'liq yakunlanmagan.\n\n"
                    f"🔄 Keyingi qadam: {next_step}\n\n"
                    f"📅 Oxirgi yangilanish: {last_updated}\n\n"
                    f"✅ Ro'yxatdan o'tishni yakunlash uchun /start ni bosing va jarayonni davom eting.\n\n"
                    f"❓ Yordam kerak bo'lsa, admin bilan bog'laning: @L19671"
                )
                
                # Send reminder message
                await context.bot.send_message(
                    chat_id=user_id,
                    text=reminder_text,
                    parse_mode='HTML'
                )
                
                sent_count += 1
                print(f"Reminder sent to user {user_id} (step: {current_step})")
                
                # Update progress (only if not the last user)
                if i < len(incomplete_users) - 1:
                    progress = int((i + 1) / len(incomplete_users) * 100)
                    progress_bar = "█" * (progress // 10) + "▱" * (10 - progress // 10)
                    
                    try:
                        await update.message.edit_text(
                            f"⏰ <b>Eslatma yuborilmoqda...</b>\n\n"
                            f"📊 Jami {len(incomplete_users)} ta to'liq bo'lmagan ro'yxatdan o'tish\n"
                            f"✅ Yuborildi: {sent_count}\n"
                            f"❌ Xato: {failed_count}\n\n"
                            f"{progress_bar} {progress}%",
                            parse_mode='HTML'
                        )
                    except Exception as edit_error:
                        print(f"Could not edit progress message: {edit_error}")
                        # Continue without updating progress
                
            except Exception as e:
                failed_count += 1
                print(f"Error sending reminder to user {user.get('user_id', 'unknown')}: {e}")

        try:
            await update.message.edit_text(
                f"✅ <b>Eslatma yuborish yakunlandi!</b>\n\n"
                f"📊 Jami: {len(incomplete_users)} ta\n"
                f"✅ Yuborildi: {sent_count} ta\n"
                f"❌ Xato: {failed_count} ta",
                parse_mode='HTML'
            )
        except Exception as edit_error:
            print(f"Could not edit final message: {edit_error}")
            # Send a new message instead
            await update.message.reply_text(
                f"✅ <b>Eslatma yuborish yakunlandi!</b>\n\n"
                f"📊 Jami: {len(incomplete_users)} ta\n"
                f"✅ Yuborildi: {sent_count} ta\n"
                f"❌ Xato: {failed_count} ta",
                parse_mode='HTML'
            )

    except Exception as e:
        await update.message.reply_text(f"❌ Export xatosi: {e}")
        print(f"Reminder error: {e}")


async def stats_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Show database statistics"""
    user_id = str(update.effective_user.id)
    if user_id != ADMIN_CHAT_ID:
        await update.message.reply_text(f"❌ Sizda ruxsat yo'q!")
        return
    
    # Reload data to get latest statistics
    load_data()
    
    complete = len([u for u in registered_users if u.get('registration_status') == 'complete'])
    incomplete = len([u for u in registered_users if u.get('registration_status') == 'incomplete'])
    old_format = len([u for u in registered_users if 'registration_status' not in u])
    
    stats_text = f"""
📊 <b>Database Statistics</b>

👥 Total Users: {len(registered_users)}
✅ Complete: {complete}
⏳ Incomplete: {incomplete}
🔄 Old Format: {old_format}

📁 Database: {DATA_FILE}
🕒 Last Updated: {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}
    """
    
    await update.message.reply_text(stats_text, parse_mode='HTML')


async def userid_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Admin command to get user info by Telegram ID. Can use /userid 123 or /123"""
    user_id = str(update.effective_user.id)
    if user_id != ADMIN_CHAT_ID:
        return  # Silently ignore for non-admins
    
    # Reload data to get latest users
    load_data()
    
    # Extract ID from command text (e.g., /123456789 -> 123456789)
    command_text = update.message.text.strip()
    
    # If it's /userid command with args
    if command_text.startswith('/userid'):
        if not context.args or len(context.args) == 0:
            await update.message.reply_text(
                "ℹ️ Foydalanish:\n"
                "• /userid <telegram_id>\n"
                "• /<telegram_id>\n\n"
                "Masalan: /userid 123456789 yoki /123456789"
            )
            return
        search_id = context.args[0]
    else:
        # Direct ID command like /123456789
        search_id = command_text[1:]  # Remove the /
        if not search_id.isdigit():
            return  # Not a valid ID command, ignore
    
    # Search for user
    found_users = [u for u in registered_users if str(u.get('user_id')) == search_id]
    
    if not found_users:
        await update.message.reply_text(f"❌ Telegram ID {search_id} topilmadi.")
        return
    
    # Show user info (take the most recent registration)
    user = found_users[-1]  # Most recent
    
    work_link = user.get('message_link', "Link yo'q")
    username = user.get('username', '')
    if not username:
        username = "yo'q"
    
    text = (
        f"👤 <b>Foydalanuvchi ma'lumotlari</b>\n\n"
        f"🆔 <b>Telegram ID:</b> <code>{user.get('user_id', '')}</code>\n"
        f"👤 <b>F.I.Sh:</b> {user.get('fullname', '')}\n"
        f"🌍 <b>Davlat:</b> {user.get('country', '')}\n"
        f"🏙️ <b>Shahar/Tuman:</b> {user.get('city', '')}\n"
        f"🎂 <b>Tug'ilgan sana:</b> {user.get('birthdate', '')}\n"
        f"📱 <b>Telefon:</b> {user.get('phone', '')}\n"
        f"📞 <b>Telegram:</b> @{username}\n"
        f"🏢 <b>Ish joyi:</b> {user.get('workplace', '')}\n"
        f"💼 <b>Mutaxassislik:</b> {user.get('specialty', '')}\n"
        f"🎓 <b>Ma'lumot:</b> {user.get('education', '')}\n"
        f"🏆 <b>Nominatsiya:</b> {user.get('nomination', '')}\n"
        f"⏰ <b>Ro'yxatdan o'tgan vaqt:</b> {user.get('registration_date', '')}\n"
        f"📎 <b>Ijodiy ish:</b> {work_link}"
    )
    
    if len(found_users) > 1:
        text += f"\n\n⚠️ <i>Bu foydalanuvchi {len(found_users)} marta ro'yxatdan o'tgan</i>"
    
    await update.message.reply_text(text, parse_mode='HTML')
    
    # Send the user's creative work file if it exists
    file_info = user.get('file')
    if file_info and file_info.get('file_id'):
        try:
            await update.message.chat.send_action(action="upload_document")
            
            if file_info.get('file_type') == 'video':
                await context.bot.send_video(
                    chat_id=update.effective_chat.id,
                    video=file_info['file_id'],
                    caption=f"🎥 Ijodiy ish: {file_info.get('file_name', 'video')}"
                )
            elif file_info.get('file_type') == 'audio':
                await context.bot.send_audio(
                    chat_id=update.effective_chat.id,
                    audio=file_info['file_id'],
                    caption=f"🎵 Ijodiy ish: {file_info.get('file_name', 'audio')}"
                )
            else:
                await context.bot.send_document(
                    chat_id=update.effective_chat.id,
                    document=file_info['file_id'],
                    caption=f"📄 Ijodiy ish: {file_info.get('file_name', 'document')}"
                )
        except Exception as e:
            await update.message.reply_text(f"⚠️ Ijodiy ishni yuklashda xatolik: {str(e)}")


async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "❌ Ro'yxatdan o'tish bekor qilindi.\n"
        "Qaytadan boshlash uchun /start buyrug'ini yuboring."
    )
    return ConversationHandler.END


def main():
    print("Bot ishga tushmoqda...")
    load_data()

    application = Application.builder().token(BOT_TOKEN).build()

    conv_handler = ConversationHandler(
        entry_points=[
            CommandHandler('start', start),
            # Before any state: button or message shows Start message
            MessageHandler(filters.ALL & ~filters.COMMAND, handle_broadcast_qatnashish),
        ],
        states={
            START_MENU: [
                # At Start menu: any message (including button) begins registration
                MessageHandler(filters.ALL & ~filters.COMMAND, startmenu_catch_all),
                CallbackQueryHandler(begin_registration_callback, pattern='^begin_reg$'),
                # Allow restart with /start
                CommandHandler('start', start),
            ],
            FULLNAME: [
                CommandHandler('start', start),
                MessageHandler(filters.TEXT & ~filters.COMMAND, fullname),
                MessageHandler(filters.ALL, fullname)
            ],
            COUNTRY: [
                CommandHandler('start', start),
                MessageHandler(filters.TEXT & ~filters.COMMAND, country),
                MessageHandler(filters.ALL, country)
            ],
            CITY: [
                CommandHandler('start', start),
                MessageHandler(filters.TEXT & ~filters.COMMAND, city),
                MessageHandler(filters.ALL, city)
            ],
            BIRTHDATE: [
                CommandHandler('start', start),
                MessageHandler(filters.TEXT & ~filters.COMMAND, birthdate),
                MessageHandler(filters.ALL, birthdate)
            ],
            PHONE: [
                CommandHandler('start', start),
                MessageHandler(filters.TEXT & ~filters.COMMAND, phone),
                MessageHandler(filters.ALL, phone)
            ],
            WORKPLACE: [
                CommandHandler('start', start),
                MessageHandler(filters.TEXT & ~filters.COMMAND, workplace),
                MessageHandler(filters.ALL, workplace)
            ],
            SPECIALTY: [
                CommandHandler('start', start),
                MessageHandler(filters.TEXT & ~filters.COMMAND, specialty),
                MessageHandler(filters.ALL, specialty)
            ],
            EDUCATION: [
                CommandHandler('start', start),
                CallbackQueryHandler(education_callback, pattern='^edu_'),
                MessageHandler(filters.ALL, education_invalid_input)
            ],
            NOMINATION: [
                CommandHandler('start', start),
                CallbackQueryHandler(nomination_callback, pattern='^nom_'),
                MessageHandler(filters.ALL, nomination_invalid_input)
            ],
            CREATIVE_WORK: [
                CommandHandler('start', start),
                MessageHandler(filters.Document.ALL | filters.VIDEO | filters.AUDIO, creative_work),
                MessageHandler(filters.ALL, creative_work)
            ],
        },
        fallbacks=[CommandHandler('cancel', cancel)],
        allow_reentry=False,
        per_chat=True,
        per_user=True,
        per_message=False,
    )

    application.add_handler(conv_handler)
    application.add_handler(CommandHandler('export', export_command))
    application.add_handler(CommandHandler('chatid', get_chat_id))
    application.add_handler(CommandHandler('broadcast', broadcast_command))
    application.add_handler(CommandHandler('reminder', reminder_command))
    application.add_handler(CommandHandler('stats', stats_command))
    application.add_handler(CommandHandler('userid', userid_command))
    # Catch any command that looks like a Telegram ID (e.g., /123456789)
    application.add_handler(MessageHandler(filters.Regex(r'^/\d+$') & filters.ChatType.PRIVATE, userid_command))
    # Handle forwarded messages from admin to recover user data
    application.add_handler(MessageHandler(filters.FORWARDED & filters.ChatType.PRIVATE, handle_forwarded_message))
    application.add_handler(CallbackQueryHandler(approve_callback, pattern='^approve_'))
    application.add_handler(CallbackQueryHandler(reject_callback, pattern='^reject_'))

    print("✅ Bot muvaffaqiyatli ishga tushdi!")
    # Schedule broadcast after startup if JobQueue is available; otherwise, use /broadcast
    if getattr(application, 'job_queue', None) is not None:
        application.job_queue.run_once(lambda ctx: application.create_task(broadcast_previous_users(application)), when=2)
    else:
        print("⚠️ JobQueue mavjud emas. Broadcastni qo'lda /broadcast orqali yuboring yoki PTB job-queue qo'shimchasini o'rnating.")
    application.run_polling(allowed_updates=Update.ALL_TYPES)


if __name__ == '__main__':
    main()
